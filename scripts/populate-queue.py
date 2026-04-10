"""
Populate CMS ContentQueue from the blog plan Excel file.

Usage: CMS_EMAIL=... CMS_PASSWORD=... python scripts/populate-queue.py [excel_path]

Environment variables:
  CMS_EMAIL     - CMS admin email (required)
  CMS_PASSWORD  - CMS admin password (required)
  CMS_URL       - CMS base URL (default: http://localhost:3001)
  EXCEL_PATH    - Path to blog plan Excel file (can also pass as first CLI argument)

Requires: openpyxl, requests
"""
import openpyxl
import requests
import json
import os
import re
import sys

CMS_URL = os.environ.get("CMS_URL", "http://localhost:3001")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "triply-pro-blog-plan-tile-installation.xlsx")

# Login to get token
def login():
    """Authenticate using environment variables CMS_EMAIL and CMS_PASSWORD."""
    email = os.environ.get("CMS_EMAIL", "")
    password = os.environ.get("CMS_PASSWORD", "")
    if not email or not password:
        print("Error: CMS_EMAIL and CMS_PASSWORD environment variables are required.")
        return None
    resp = requests.post(f"{CMS_URL}/api/users/login", json={"email": email, "password": password})
    if resp.status_code == 200:
        data = resp.json()
        print(f"Logged in as {email}")
        return data.get("token")
    print(f"Could not login as {email} (status {resp.status_code}).")
    return None

def to_slug(keyword):
    return re.sub(r'(^-|-$)', '', re.sub(r'[^a-z0-9]+', '-', keyword.lower()))

def map_tier(tier):
    """Map Excel tier to CMS articleType"""
    mapping = {"HUB": "hub", "SUB-PILLAR": "pillar", "SPOKE": "spoke"}
    return mapping.get(tier, "spoke")

def map_priority(p):
    """Map P0-P3 to S1-S3"""
    mapping = {"P0": "S1", "P1": "S1", "P2": "S2", "P3": "S3"}
    return mapping.get(p, "S2")

def parse_target_words(tw):
    """Parse '4,000–5,000' → midpoint integer"""
    if not tw:
        return 1500
    tw = str(tw)
    # Handle various dash types and comma-separated numbers
    nums = re.findall(r'[\d,]+', tw)
    if len(nums) >= 2:
        low = int(nums[0].replace(',', ''))
        high = int(nums[1].replace(',', ''))
        return (low + high) // 2
    elif len(nums) == 1:
        return int(nums[0].replace(',', ''))
    return 1500

def infer_service_type(keyword, parent_hub):
    """Infer serviceType from keyword and parent context"""
    kw = keyword.lower()
    ph = (parent_hub or "").lower()

    # Check keyword first for specific service mentions
    if any(w in kw for w in ['bathroom', 'shower', 'toilet', 'vanity']):
        if 'shower' in kw:
            return 'shower'
        return 'bathroom'
    if any(w in kw for w in ['floor', 'subfloor', 'concrete', 'plywood', 'wood floor']):
        return 'floor'
    if any(w in kw for w in ['backsplash', 'kitchen', 'cabinet']):
        return 'backsplash'
    if any(w in kw for w in ['repair', 'fix', 'crack', 'replace']):
        return 'repair'
    if any(w in kw for w in ['reglaz', 'refinish']):
        return 'reglazing'

    # Check parent hub for context
    if 'bathroom' in ph:
        return 'bathroom'
    if 'floor' in ph:
        return 'floor'
    if 'backsplash' in ph:
        return 'backsplash'

    return 'general'

def infer_article_style(keyword, tier, notes):
    """Infer articleStyle from content"""
    kw = keyword.lower()
    notes_lower = (notes or "").lower()

    if tier == "HUB":
        return "standard"
    if any(w in kw for w in ['how to', 'how-to', 'diy', 'steps', 'process', 'guide', 'install ']):
        return "how-to"
    if any(w in kw for w in ['cost', 'price', 'rate', 'calculator', 'estimate', 'budget', 'quote']):
        return "data-heavy"
    if any(w in kw for w in ['pattern', 'design', 'ideas', 'layout', 'types', 'options']):
        return "listicle"
    if any(w in kw for w in ['vs', 'comparison', 'compare']):
        return "comparison"
    return "standard"

def get_hub_slug(parent_hub, tier):
    """All items belong to the 'tile-installation' hub"""
    if tier == "HUB":
        return None
    return "tile-installation"

def main():
    token = login()
    if not token:
        print("Failed to authenticate. Please check CMS credentials.")
        sys.exit(1)

    headers = {"Authorization": f"JWT {token}", "Content-Type": "application/json"}

    # Check existing queue items to avoid duplicates
    resp = requests.get(f"{CMS_URL}/api/content-queue?limit=500", headers=headers)
    existing_slugs = set()
    if resp.status_code == 200:
        for doc in resp.json().get("docs", []):
            existing_slugs.add(doc.get("slug"))
    print(f"Found {len(existing_slugs)} existing queue items")

    # Determine Excel path: CLI arg > env var > default
    excel_path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH

    # Read Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Content Dashboard']

    # Also read outlines for hub and sub-pillars
    hub_outlines = {}
    ws_hub = wb['Hub Outline']
    hub_outline_items = []
    for row in ws_hub.iter_rows(min_row=2, max_row=ws_hub.max_row, values_only=True):
        if row[0] is not None:
            hub_outline_items.append({
                "order": int(row[0]),
                "anchorId": row[1] or "",
                "heading": row[2] or "",
                "summary": row[3] or "",
                "linksTo": to_slug(row[4]) if row[4] else "",
            })
    hub_outlines["tile-installation"] = hub_outline_items

    sub_pillar_outlines = {}
    ws_sp = wb['Sub-Pillar Outlines']
    for row in ws_sp.iter_rows(min_row=2, max_row=ws_sp.max_row, values_only=True):
        if row[0] is not None:
            sp_slug = to_slug(str(row[0]))
            if sp_slug not in sub_pillar_outlines:
                sub_pillar_outlines[sp_slug] = []
            sub_pillar_outlines[sp_slug].append({
                "order": int(row[1]) if row[1] else len(sub_pillar_outlines[sp_slug]) + 1,
                "anchorId": "",
                "heading": row[2] or "",
                "summary": row[3] or "",
                "linksTo": "",
            })

    # Process rows
    created = 0
    skipped = 0
    errors = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[2]:  # no keyword
            continue

        num = row[0]
        tier = row[1]
        keyword = row[2]
        niche = row[3]
        parent_hub = row[4]
        search_vol = row[5]
        seo_diff = row[6]
        cpc = row[7]
        target_words = row[8]
        h2_count = row[9]
        status = row[10]
        priority = row[11]
        notes = row[12]

        slug = to_slug(keyword)

        if slug in existing_slugs:
            skipped += 1
            continue

        # Build the queue item
        item = {
            "keyword": keyword,
            "slug": slug,
            "articleType": map_tier(tier),
            "serviceType": infer_service_type(keyword, parent_hub),
            "articleStyle": infer_article_style(keyword, tier, notes),
            "priority": map_priority(priority),
            "status": "queued",
            "searchVolume": int(search_vol) if search_vol else 0,
            "seoDifficulty": int(seo_diff) if seo_diff else 0,
            "targetWords": parse_target_words(target_words),
            "notes": notes or "",
        }

        # Set parent/hub slugs
        if tier != "HUB":
            item["hubSlug"] = "tile-installation"
        if parent_hub and str(parent_hub).strip() and parent_hub != "\u2014":
            parent_slug = to_slug(str(parent_hub))
            if parent_slug:
                item["parentSlug"] = parent_slug

        # Add outlines for hub and sub-pillars
        if tier == "HUB" and slug in hub_outlines:
            item["outline"] = hub_outlines[slug]
        elif tier == "SUB-PILLAR" and slug in sub_pillar_outlines:
            item["outline"] = sub_pillar_outlines[slug]

        # Create in CMS
        resp = requests.post(f"{CMS_URL}/api/content-queue", headers=headers, json=item)
        if resp.status_code in (200, 201):
            created += 1
            existing_slugs.add(slug)
            tier_label = f"[{tier:11s}]"
            print(f"  {created:3d}. {tier_label} {keyword}")
        else:
            errors += 1
            err = resp.json() if resp.headers.get('content-type', '').startswith('application/json') else resp.text
            print(f"  ERROR: {keyword} → {err}")

    print(f"\nDone! Created: {created}, Skipped (existing): {skipped}, Errors: {errors}")

if __name__ == "__main__":
    main()
